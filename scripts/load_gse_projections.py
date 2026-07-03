#!/usr/bin/env python3
"""Extract data/gse_load_2026_2030.json from the committed
``load 2026-2030.rar``.

Rerun this after replacing the .rar (or the underlying xlsx set) if the
GSE forecast is refreshed. The script prefers ``unrar-free`` (installed
by ``apt install unrar-free``); if that is missing it falls back to the
Python ``rarfile`` module (``pip install rarfile``) which requires the
``unrar`` binary to be present system-wide.

The output JSON is ~340 KiB — small enough to commit; the .rar itself
is kept in the repo so this script is a one-command refresh.

Usage:
    python scripts/load_gse_projections.py
    python scripts/load_gse_projections.py --rar load\\ 2026-2030.rar \\
                                            --out data/gse_load_2026_2030.json
"""
from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

DEFAULT_RAR = "load 2026-2030.rar"
DEFAULT_OUT = "data/gse_load_2026_2030.json"
YEARS = (2026, 2027, 2028, 2029, 2030)


def _extract_rar(rar: Path, dest: Path) -> Path:
    """Extract ``rar`` into ``dest`` and return the directory containing
    the xlsx files. Uses whichever unrar tool is on PATH."""
    dest.mkdir(parents=True, exist_ok=True)
    for tool in ("unrar-free", "unrar", "unar", "7z"):
        if shutil.which(tool):
            if tool == "7z":
                cmd = [tool, "x", "-y", f"-o{dest}", str(rar)]
            elif tool == "unar":
                cmd = [tool, "-force-overwrite", "-o", str(dest), str(rar)]
            else:
                cmd = [tool, "x", str(rar), str(dest) + "/"]
            proc = subprocess.run(cmd, capture_output=True, text=True)
            if proc.returncode == 0:
                break
    else:
        # Try python rarfile
        try:
            import rarfile          # type: ignore
        except ImportError as e:
            raise RuntimeError(
                "no unrar binary found (tried unrar-free / unrar / unar / 7z) "
                "and python-rarfile is not installed"
            ) from e
        with rarfile.RarFile(str(rar)) as rf:
            rf.extractall(str(dest))
    # Locate the xlsx dir (rar puts them under 'load 2026-2030/' typically).
    for entry in dest.rglob("GSE_PLEXOS_sc2_2026.xlsx"):
        return entry.parent
    raise RuntimeError(f"extraction succeeded but no xlsx found under {dest}")


def _load_year_xlsx(xlsx: Path, year: int) -> tuple[list[float], float | None]:
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx), data_only=True, read_only=True)
    try:
        sheet_name = f"PLEXOS_sc2_{year}"
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        target = None
        vals: list[float] = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if i == 0 and len(row) >= 5 and row[4] is not None:
                try: target = float(row[4])
                except (TypeError, ValueError): target = None
            if len(row) < 2 or row[1] is None:
                break
            vals.append(round(float(row[1]), 3))
        return vals, target
    finally:
        wb.close()


def main() -> int:
    ap = argparse.ArgumentParser(description="Extract GSE hourly-load JSON snapshot")
    ap.add_argument("--rar", default=DEFAULT_RAR)
    ap.add_argument("--out", default=DEFAULT_OUT)
    args = ap.parse_args()

    rar = Path(args.rar)
    if not rar.exists():
        print(f"❌ rar not found: {rar}", file=sys.stderr); return 1

    with tempfile.TemporaryDirectory() as td:
        xdir = _extract_rar(rar, Path(td))
        snap = {
            "meta": {
                "source": "GSE PLEXOS sc2 hourly load projections (P50 central) 2026-2030",
                "scenario": "P50 — ცენ.",
                "unit": "MW",
                "resolution": "hourly (8760 per year)",
                "target_gwh_per_year": {},
                "peak_mw_per_year": {},
                "notes": "Extracted from 'load 2026-2030.rar'. Refresh via scripts/load_gse_projections.py.",
            },
            "load_mw_by_year": {},
        }
        for y in YEARS:
            xlsx = xdir / f"GSE_PLEXOS_sc2_{y}.xlsx"
            if not xlsx.exists():
                print(f"❌ missing {xlsx}", file=sys.stderr); return 1
            vals, target = _load_year_xlsx(xlsx, y)
            snap["load_mw_by_year"][str(y)] = vals
            snap["meta"]["target_gwh_per_year"][str(y)] = target
            snap["meta"]["peak_mw_per_year"][str(y)] = round(max(vals), 1)
            print(f"  {y}: {len(vals)}h sum={sum(vals)/1000:.1f} GWh "
                  f"peak={max(vals):.1f} MW target={target}")

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(snap, ensure_ascii=False, separators=(",", ":")))
    print(f"\n✅ wrote {out} — {out.stat().st_size/1024:.1f} KiB")
    return 0


if __name__ == "__main__":
    sys.exit(main())
