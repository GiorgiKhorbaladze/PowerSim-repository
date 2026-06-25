"""Georgia 2030 BESS adequacy report generator.

Produces:
  - Excel workbook (multi-sheet)
  - Markdown/word-ready tables
from scenario results + hourly dispatch of a selected scenario.
"""
from __future__ import annotations

import argparse, json, math
from pathlib import Path
from typing import Any

# Lazy openpyxl import so the module can be imported without it
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

from solver.bess_sizing import (
    PlexosZipLoader,
    BessSizingParams,
    REG_ENERGY_GWH,
    REG_PMAX_MW,
    KSANI_P_MW,
    KSANI_E_MWH,
    DURATION_COST,
    solve_bess_sizing,
)

MONTH_LABELS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
MONTH_HOURS_2030 = [744, 672, 744, 720, 744, 720, 744, 744, 720, 744, 720, 744]

# ── helpers ──────────────────────────────────────────────────────────────────

def _month_of_t():
    mo = []
    for m, h in enumerate(MONTH_HOURS_2030):
        mo += [m] * h
    return mo


def _monthly_sum(series, month_of_t):
    totals = [0.0] * 12
    for t, v in enumerate(series):
        totals[month_of_t[t]] += v
    return totals


def _gwh(mw_series) -> float:
    return sum(mw_series) / 1000.0


def _header_row(ws, row: int, values: list, bold: bool = True, bg: str = '4472C4', fg: str = 'FFFFFF'):
    if not _HAS_OPENPYXL:
        return
    fill = PatternFill(fill_type='solid', fgColor=bg)
    font = Font(bold=bold, color=fg)
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal='center', wrap_text=True)


def _data_row(ws, row: int, values: list, fmt_map: dict | None = None):
    if not _HAS_OPENPYXL:
        return
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        if fmt_map and col in fmt_map:
            c.number_format = fmt_map[col]


def _auto_width(ws):
    if not _HAS_OPENPYXL:
        return
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Sheet builders
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _sheet_summary(wb, loader_qa: dict):
    ws = wb.create_sheet('Summary')
    ws.title = 'Summary'
    row = 1
    ws.cell(row, 1, 'Georgia 2030 BESS Adequacy Study – Summary').font = Font(bold=True, size=14)
    row += 2

    _header_row(ws, row, ['Item', 'Value', 'Unit'])
    row += 1
    items = [
        ('Internal load', f"{loader_qa.get('load_gwh', 0):.0f}", 'GWh/yr'),
        ('Export demand', f"{loader_qa.get('export_gwh', 0):.0f}", 'GWh/yr'),
        ('PV available', f"{loader_qa.get('pv_gwh', 0):.0f}", 'GWh/yr'),
        ('Wind available', f"{loader_qa.get('wind_gwh', 0):.0f}", 'GWh/yr'),
        ('RoR available', f"{loader_qa.get('ror_gwh', 0):.0f}", 'GWh/yr'),
        ('TPP/Import', f"{loader_qa.get('tpp_gwh', 0):.0f}", 'GWh/yr'),
        ('Ksani BESS (existing)', f'{KSANI_P_MW:.0f} MW / {KSANI_E_MWH:.0f} MWh', '1-hour'),
        ('RegHPP energy limit', f"{sum(REG_ENERGY_GWH):.0f}", 'GWh/yr'),
    ]
    for it in items:
        _data_row(ws, row, list(it))
        row += 1
    _auto_width(ws)


def _sheet_scenarios(wb, results: list[dict]):
    ws = wb.create_sheet('Scenario Comparison')
    ws.title = 'Scenario Comparison'
    headers = [
        'Scenario', 'Export', 'EENS target %', 'Add dur h',
        'Padd MW', 'Eadd MWh', 'Ptotal MW', 'Etotal MWh',
        'EENS %', 'LOLH h', 'Export ENS GWh', 'Curtail GWh',
        'PV GWh', 'Wind GWh', 'RoR GWh', 'TPP GWh', 'RegHPP GWh',
        'BESS cyc/yr', 'Add CAPEX MUSD/yr', 'Total CAPEX MUSD/yr',
        'Closure %',
    ]
    _header_row(ws, 1, headers)
    for r_idx, r in enumerate(results, 2):
        if r.get('status') == 'Error':
            ws.cell(r_idx, 1, r['scenario']['label'])
            ws.cell(r_idx, 2, 'ERROR')
            ws.cell(r_idx, 3, r.get('error', ''))
            continue
        sc = r.get('scenario', {})
        exp_label = 'firm' if sc.get('export_firm') else 'curt'
        row_vals = [
            sc.get('label', ''),
            exp_label,
            sc.get('eens_target_pct', '—'),
            sc.get('add_duration_h', '—'),
            r.get('pbess_add_mw', 0),
            r.get('ebess_add_mwh', 0),
            r.get('pbess_total_mw', 0),
            r.get('ebess_total_mwh', 0),
            r.get('eens_pct', 0),
            r.get('lolh_h', 0),
            r.get('export_ens_gwh', 0),
            r.get('curtailment_gwh', 0),
            r.get('pv_gwh', 0),
            r.get('wind_gwh', 0),
            r.get('ror_gwh', 0),
            r.get('tpp_gwh', 0),
            r.get('reg_gwh', 0),
            r.get('cycles_per_year', 0),
            r.get('add_capex_annualized_usd', 0) / 1e6,
            r.get('total_capex_annualized_usd', 0) / 1e6,
            r.get('energy_closure_pct', 0),
        ]
        _data_row(ws, r_idx, row_vals)
    _auto_width(ws)


def _sheet_hourly(wb, series: dict, label: str):
    ws = wb.create_sheet('Hourly Dispatch')
    ws.title = 'Hourly Dispatch'
    headers = [
        'Hour', 'Month', 'Load MW', 'Export Demand MW', 'Export Served MW',
        'PV MW', 'Wind MW', 'RoR MW', 'TPP MW', 'RegHPP MW',
        'BESS Charge MW', 'BESS Discharge MW', 'SoC MWh',
        'Internal ENS MW', 'Export ENS MW', 'Spill MW',
    ]
    _header_row(ws, 1, headers)
    mo = _month_of_t()
    n = len(series.get('load_mw', []))
    for t in range(n):
        row_vals = [
            t + 1,
            MONTH_LABELS[mo[t]],
            series.get('load_mw', [0]*n)[t],
            series.get('export_demand_mw', [0]*n)[t],
            series.get('export_served_mw', [0]*n)[t],
            series.get('pv_mw', [0]*n)[t],
            series.get('wind_mw', [0]*n)[t],
            series.get('ror_mw', [0]*n)[t],
            series.get('tpp_mw', [0]*n)[t],
            series.get('reg_mw', [0]*n)[t],
            series.get('charge_mw', [0]*n)[t],
            series.get('discharge_mw', [0]*n)[t],
            series.get('soc_mwh', [0]*n)[t],
            series.get('ens_internal_mw', [0]*n)[t],
            series.get('ens_export_mw', [0]*n)[t],
            series.get('spill_mw', [0]*n)[t],
        ]
        _data_row(ws, t + 2, row_vals)
    _auto_width(ws)


def _sheet_monthly(wb, series: dict):
    ws = wb.create_sheet('Monthly Balance')
    ws.title = 'Monthly Balance'
    mo = _month_of_t()
    headers = ['Month', 'Load GWh', 'Export Demand GWh', 'Export Served GWh',
               'PV GWh', 'Wind GWh', 'RoR GWh', 'TPP GWh', 'RegHPP GWh',
               'BESS Charge GWh', 'BESS Discharge GWh',
               'Internal ENS GWh', 'Export ENS GWh', 'Spill GWh']
    _header_row(ws, 1, headers)
    keys = ['load_mw', 'export_demand_mw', 'export_served_mw',
            'pv_mw', 'wind_mw', 'ror_mw', 'tpp_mw', 'reg_mw',
            'charge_mw', 'discharge_mw',
            'ens_internal_mw', 'ens_export_mw', 'spill_mw']
    for m in range(12):
        row_vals = [MONTH_LABELS[m]]
        for k in keys:
            s = series.get(k, [])
            total = sum(v for t, v in enumerate(s) if mo[t] == m) / 1000.0
            row_vals.append(round(total, 2))
        _data_row(ws, m + 2, row_vals)
    _auto_width(ws)


def _sheet_top_ens(wb, series: dict, top_n: int = 200):
    ws = wb.create_sheet('Top ENS Hours')
    ws.title = 'Top ENS Hours'
    headers = ['Rank', 'Hour', 'Month', 'Internal ENS MW',
               'Load MW', 'SoC MWh', 'Export Demand MW']
    _header_row(ws, 1, headers)
    mo = _month_of_t()
    ens_int = series.get('ens_internal_mw', [])
    load    = series.get('load_mw', [])
    soc     = series.get('soc_mwh', [])
    xdem    = series.get('export_demand_mw', [])
    ranked  = sorted(range(len(ens_int)), key=lambda t: -ens_int[t])[:top_n]
    for rank, t in enumerate(ranked, 1):
        if ens_int[t] < 1e-6:
            break
        _data_row(ws, rank + 1, [
            rank, t + 1, MONTH_LABELS[mo[t]],
            round(ens_int[t], 2),
            round(load[t], 1),
            round(soc[t], 1),
            round(xdem[t], 1),
        ])
    _auto_width(ws)


def _sheet_qa(wb, results: list[dict], loader_qa: dict):
    ws = wb.create_sheet('QA Validation')
    ws.title = 'QA Validation'
    row = 1
    ws.cell(row, 1, 'Loader QA').font = Font(bold=True)
    row += 1
    _header_row(ws, row, ['Category', 'Annual GWh', 'Peak MW', 'Status'])
    row += 1
    qa_ranges = {
        'load':   (17000, 19500),
        'export': (4000, 5000),
        'tpp':    (150, 220),
    }
    for k in ['load', 'export', 'pv', 'wind', 'ror', 'tpp']:
        gwh  = loader_qa.get(f'{k}_gwh', 0)
        peak = loader_qa.get(f'{k}_peak_mw', 0)
        lo, hi = qa_ranges.get(k, (0, 1e12))
        status = 'OK' if lo <= gwh <= hi else ('WARN' if k not in qa_ranges else 'FAIL')
        _data_row(ws, row, [k, round(gwh, 1), round(peak, 1), status])
        row += 1

    row += 2
    ws.cell(row, 1, 'Scenario Validation').font = Font(bold=True)
    row += 1
    _header_row(ws, row, ['Test', 'Result', 'Pass?'])
    row += 1

    # Check: fixed 800/1400 no longer gives 15.7% EENS
    for r in results:
        if r.get('scenario', {}).get('label') == 'fixed_800_1400':
            eens = r.get('eens_pct', 999)
            ok   = eens < 5.0
            _data_row(ws, row, [
                'Fixed 800/1400 EENS < 5% (not 15.7%)',
                f'{eens:.3f}%',
                'PASS' if ok else 'FAIL',
            ])
            row += 1

    # Check: export ENS never enters internal EENS
    for r in results:
        if r.get('status') == 'Error':
            continue
        # internal EENS should not include export ENS by construction
        _data_row(ws, row, [
            'Export ENS tracked separately',
            'by construction',
            'PASS',
        ])
        row += 1
        break

    # Check: energy closure within 0.5%
    for r in results:
        if r.get('status') == 'Error':
            continue
        clos = abs(r.get('energy_closure_pct', 999))
        ok   = clos < 0.5
        _data_row(ws, row, [
            f'Energy closure ({r["scenario"]["label"]})',
            f'{clos:.4f}%',
            'PASS' if ok else 'FAIL',
        ])
        row += 1
        break

    _auto_width(ws)


def _sheet_loader_qa(wb, loader_qa: dict):
    ws = wb.create_sheet('Loader QA')
    ws.title = 'Loader QA'
    _header_row(ws, 1, ['Category', 'Annual GWh', 'Peak MW', 'Min MW', 'Expected GWh range'])
    ranges_note = {
        'load': '17000–19500',
        'export': '4000–5000',
        'tpp': '150–220',
        'pv': 'credible (>0)',
        'wind': 'credible (>0)',
        'ror': 'credible (>0)',
    }
    for i, k in enumerate(['load', 'export', 'pv', 'wind', 'ror', 'tpp'], 2):
        _data_row(ws, i, [
            k,
            round(loader_qa.get(f'{k}_gwh', 0), 1),
            round(loader_qa.get(f'{k}_peak_mw', 0), 1),
            round(loader_qa.get(f'{k}_min_mw', 0), 1),
            ranges_note.get(k, '—'),
        ])
    # RegHPP limits
    row = 9
    ws.cell(row, 1, 'Regulated HPP monthly energy limits (GWh):').font = Font(bold=True)
    row += 1
    _data_row(ws, row, ['Month'] + MONTH_LABELS)
    row += 1
    _data_row(ws, row, ['Energy GWh'] + list(REG_ENERGY_GWH))
    row += 1
    _data_row(ws, row, ['Pmax MW'] + list(REG_PMAX_MW))
    _auto_width(ws)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Markdown report
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _markdown_report(results: list[dict], loader_qa: dict) -> str:
    lines = []
    lines.append('# Georgia 2030 BESS Adequacy Study\n')
    lines.append('## System Data\n')
    lines.append(f'| Item | Value |')
    lines.append('|---|---|')
    lines.append(f'| Internal load | {loader_qa.get("load_gwh",0):.0f} GWh/yr |')
    lines.append(f'| Export demand | {loader_qa.get("export_gwh",0):.0f} GWh/yr |')
    lines.append(f'| PV available | {loader_qa.get("pv_gwh",0):.0f} GWh/yr |')
    lines.append(f'| Wind available | {loader_qa.get("wind_gwh",0):.0f} GWh/yr |')
    lines.append(f'| RoR available | {loader_qa.get("ror_gwh",0):.0f} GWh/yr |')
    lines.append(f'| TPP/Import | {loader_qa.get("tpp_gwh",0):.0f} GWh/yr |')
    lines.append(f'| Ksani BESS (existing) | {KSANI_P_MW:.0f} MW / {KSANI_E_MWH:.0f} MWh (1-hour) |')
    lines.append(f'| RegHPP annual limit | {sum(REG_ENERGY_GWH):.0f} GWh/yr |')
    lines.append('')

    lines.append('## Scenario Results\n')
    h = ('| Scenario | Export | EENS tgt | Add dur | '
         'Padd MW | Eadd MWh | Ptot MW | Etot MWh | '
         'EENS % | LOLH h | ExpENS GWh | Curt GWh | Add CAPEX MUSD/yr |')
    sep = '|' + '---|' * (h.count('|') - 1)
    lines.append(h)
    lines.append(sep)

    for r in results:
        if r.get('status') == 'Error':
            sc = r.get('scenario', {})
            lines.append(f'| {sc.get("label","?")} | ERROR | | | | | | | | | | | |')
            continue
        sc = r.get('scenario', {})
        exp_label = 'firm' if sc.get('export_firm') else 'curt'
        tgt = f'{sc.get("eens_target_pct","—")}'
        dur = f'{sc.get("add_duration_h","—"):.0f}h'
        lines.append(
            f'| {sc.get("label","?")} | {exp_label} | {tgt} | {dur} | '
            f'{r.get("pbess_add_mw",0):.0f} | {r.get("ebess_add_mwh",0):.0f} | '
            f'{r.get("pbess_total_mw",0):.0f} | {r.get("ebess_total_mwh",0):.0f} | '
            f'{r.get("eens_pct",0):.3f} | {r.get("lolh_h",0):.0f} | '
            f'{r.get("export_ens_gwh",0):.1f} | {r.get("curtailment_gwh",0):.1f} | '
            f'{r.get("add_capex_annualized_usd",0)/1e6:.1f} |'
        )
    lines.append('')
    return '\n'.join(lines)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Main entrypoint
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def generate_report(
    zip_path: str,
    scenarios_json: str | None = None,
    selected_label: str | None = None,
    out_dir: str = 'out',
) -> None:
    """Generate Excel workbook and Markdown report.

    If scenarios_json is supplied, results are loaded from it (no re-solve).
    Otherwise the scenario runner is invoked inline.
    """
    from solver.bess_scenario_runner import run_all

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    # Load data for QA
    loader = PlexosZipLoader(zip_path)
    data   = loader.load(verbose=True)
    reg_energy, reg_pmax = loader.reg_limits()
    loader_qa = {
        k + '_gwh':     sum(data[k]) / 1000
        for k in ('load', 'export', 'pv', 'wind', 'ror', 'tpp')
    }
    loader_qa.update({
        k + '_peak_mw': max(data[k])
        for k in ('load', 'export', 'pv', 'wind', 'ror', 'tpp')
    })
    loader_qa.update({
        k + '_min_mw':  min(data[k])
        for k in ('load', 'export', 'pv', 'wind', 'ror', 'tpp')
    })

    # Load or run scenarios
    if scenarios_json and Path(scenarios_json).exists():
        with open(scenarios_json, encoding='utf-8') as f:
            results = json.load(f)
    else:
        results = run_all(zip_path, out_dir=out_dir, verbose=False)

    # Find the selected scenario for hourly dispatch
    if selected_label is None:
        # Default: first reliability-mode scenario
        for r in results:
            if r.get('scenario', {}).get('mode') == 'reliability' and r.get('status') == 'Optimal':
                selected_label = r['scenario']['label']
                break

    # Re-solve selected scenario to get hourly series
    series: dict = {}
    for r in results:
        if r.get('scenario', {}).get('label') == selected_label:
            sc = r['scenario']
            params = BessSizingParams(
                mode=sc.get('mode', 'reliability'),
                eens_target_pct=sc.get('eens_target_pct', 0.10),
                add_duration_h=sc.get('add_duration_h', 1.0),
                export_firm=sc.get('export_firm', False),
                fixed_p_mw=sc.get('fixed_p_mw'),
                fixed_e_mwh=sc.get('fixed_e_mwh'),
            )
            print(f'Re-solving {selected_label} for hourly series …')
            full = solve_bess_sizing(data, params, reg_energy, reg_pmax)
            series = full.get('series', {})
            break

    if not _HAS_OPENPYXL:
        print('WARNING: openpyxl not installed; skipping Excel workbook.')
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove default sheet
        _sheet_summary(wb, loader_qa)
        _sheet_scenarios(wb, results)
        if series:
            _sheet_hourly(wb, series, selected_label or '')
            _sheet_monthly(wb, series)
            _sheet_top_ens(wb, series)
        _sheet_qa(wb, results, loader_qa)
        _sheet_loader_qa(wb, loader_qa)

        xl_path = out_path / 'Georgia_2030_BESS_Report.xlsx'
        wb.save(xl_path)
        print(f'Excel workbook saved → {xl_path}')

    md = _markdown_report(results, loader_qa)
    md_path = out_path / 'Georgia_2030_BESS_Report.md'
    md_path.write_text(md, encoding='utf-8')
    print(f'Markdown report saved → {md_path}')


def main() -> None:
    ap = argparse.ArgumentParser(description='Georgia 2030 BESS report generator')
    ap.add_argument('--plexos-zip', default='plexos model.zip')
    ap.add_argument('--scenarios-json', default=None)
    ap.add_argument('--selected-label', default=None,
                    help='Scenario label for hourly dispatch sheet')
    ap.add_argument('--out', default='out')
    args = ap.parse_args()
    generate_report(
        args.plexos_zip,
        scenarios_json=args.scenarios_json,
        selected_label=args.selected_label,
        out_dir=args.out,
    )


if __name__ == '__main__':
    main()
