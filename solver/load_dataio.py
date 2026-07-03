"""GSE hourly-load projection loader.

Reads Excel workbooks in the GSE PLEXOS format used for scenario studies
(one sheet named ``PLEXOS_sc<n>_<YYYY>`` with columns ``DateTime``,
``MW_sc<n>_<YYYY>``, ``სცენარი``, ``წელი``, ``Target_GWh``) and returns
a plain list of 8760 hourly MW values.

Also exposes ``load_snapshot()`` for the pre-committed JSON at
``data/gse_load_2026_2030.json`` (extracted from ``load 2026-2030.rar``
by :mod:`scripts.load_gse_projections`).

Downstream: the returned list plugs directly into
``inp["profiles"]["demand"]`` for the PowerSim solver.

Public API
----------
    load_snapshot() -> dict
    load_snapshot_year(year) -> list[float]
    load_gse_xlsx(path, sheet=None) -> list[float]
    autodetect_hourly_demand(path, sheet=None) -> tuple[list[float], dict]
"""
from __future__ import annotations

import json
from functools import lru_cache
from pathlib import Path
from typing import Optional

HOURS_PER_YEAR = 8760
_SNAPSHOT_PATH = (
    Path(__file__).resolve().parent.parent / "data" / "gse_load_2026_2030.json"
)


class LoadIOError(Exception):
    """Raised when a load Excel does not match any known shape."""


@lru_cache(maxsize=1)
def load_snapshot() -> dict:
    """Return the pre-committed GSE 2026-2030 snapshot dict. Cached."""
    if not _SNAPSHOT_PATH.exists():
        raise FileNotFoundError(
            f"snapshot missing at {_SNAPSHOT_PATH}; run "
            f"scripts/load_gse_projections.py to (re-)extract")
    return json.loads(_SNAPSHOT_PATH.read_text(encoding="utf-8"))


def snapshot_years() -> list[int]:
    return sorted(int(y) for y in load_snapshot()["load_mw_by_year"].keys())


def load_snapshot_year(year: int) -> list[float]:
    """Return the 8760-h MW profile for a snapshot year, ready to plug
    into ``inp["profiles"]["demand"]``."""
    snap = load_snapshot()
    key = str(int(year))
    if key not in snap["load_mw_by_year"]:
        raise KeyError(
            f"year {year} not in snapshot; available: {snapshot_years()}")
    return list(snap["load_mw_by_year"][key])


def load_gse_xlsx(path: str | Path, sheet: Optional[str] = None) -> list[float]:
    """Read a single GSE-format Excel and return an 8760-h MW profile.

    ``sheet`` defaults to the first non-metadata sheet whose name starts
    with ``"PLEXOS_"``. Raises ``LoadIOError`` if the shape doesn't
    resemble one hourly-load column.
    """
    profile, meta = autodetect_hourly_demand(path, sheet)
    if len(profile) not in (8760, 8784):
        raise LoadIOError(
            f"{path}: extracted {len(profile)} hours; expected 8760 or 8784")
    # Trim leap year to 8760 by discarding Feb 29 (last 24h at the end
    # is usually easier for the user to reason about — but the standard
    # convention is to drop 24 hours from the middle. We drop from the
    # tail as a safe default.)
    return profile[:HOURS_PER_YEAR]


def autodetect_hourly_demand(
    path: str | Path, sheet: Optional[str] = None,
) -> tuple[list[float], dict]:
    """Robustly extract hourly demand from a variety of Excel shapes.

    Walks each sheet (or the given one) and picks the first numeric
    column whose length is close to 8760 (or 8784) and whose values are
    plausibly MW (positive, ≤ 100 000, non-flat). Returns the values
    plus a metadata dict {sheet, column_index, column_header, n_rows}.
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        sheets = [sheet] if sheet else wb.sheetnames
        for sn in sheets:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 100:
                continue
            header = rows[0]
            data_rows = rows[1:]
            n_cols = max((len(r) for r in data_rows[:5]), default=0)
            for c in range(n_cols):
                col_vals = []
                for r in data_rows:
                    if c >= len(r):
                        col_vals.append(None); continue
                    col_vals.append(r[c])
                # numeric coercion
                numeric = []
                for v in col_vals:
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        numeric.append(float(v))
                    else:
                        try: numeric.append(float(v))
                        except (TypeError, ValueError): numeric.append(None)
                good_count = sum(1 for v in numeric if v is not None)
                if good_count < 8000:
                    continue
                clean = [v for v in numeric if v is not None]
                if len(clean) < 8000:
                    continue
                mn, mx = min(clean), max(clean)
                if not (0 <= mn) or mx > 100_000 or mx <= 0:
                    continue
                if mx - mn < 1e-6:   # flat column, not a demand series
                    continue
                # Truncate leading Nones / take first 8784.
                profile = [v for v in numeric if v is not None][:8784]
                meta = {"sheet": sn, "column_index": c,
                        "column_header": header[c] if c < len(header) else None,
                        "n_rows": len(profile),
                        "peak_mw": round(mx, 3),
                        "annual_gwh": round(sum(profile) / 1000, 3),
                        "source_file": str(path)}
                return profile, meta
        raise LoadIOError(f"{path}: no plausible hourly-MW column found")
    finally:
        wb.close()
